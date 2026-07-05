"""
Utility to parse uploaded PDF, Excel (xlsx), and JSON files into clean text structures.
"""
from __future__ import annotations

import io
import json
import logging
from pypdf import PdfReader
import openpyxl

logger = logging.getLogger(__name__)


def parse_pdf(file_bytes: bytes) -> str:
    """Extract text page-by-page from a PDF byte stream."""
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        text_parts = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                text_parts.append(f"--- Page {i+1} ---\n{text}")
        return "\n\n".join(text_parts)
    except Exception as e:
        logger.error("Failed to parse PDF: %s", e)
        return ""


def parse_excel(file_bytes: bytes) -> str:
    """Extract rows and columns from Excel sheets as pipe-separated lines of text."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        logger.error("Failed to parse Excel: %s", e)
        return ""


def parse_file(filename: str, file_bytes: bytes) -> str:
    """Dispatches parser depending on the file extension."""
    ext = filename.split(".")[-1].lower()
    if ext == "pdf":
        return parse_pdf(file_bytes)
    elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
        return parse_excel(file_bytes)
    elif ext == "json":
        try:
            data = json.loads(file_bytes.decode("utf-8"))
            return json.dumps(data, indent=2)
        except Exception:
            return file_bytes.decode("utf-8", errors="ignore")
    else:
        return file_bytes.decode("utf-8", errors="ignore")
